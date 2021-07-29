# imports
from typing import Text
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import pandas as pd
import time

# Config Inicial
WEB_PAGE = "http://scw.pjn.gov.ar/scw/home.seam"
BUSQUEDA_ID = "formPublica:porParte:header:inactive"
VALOR_BUSQUEDA = "SOCIEDAD"
SELECT_ID = "formPublica:camaraPartes"
JURISDICCION_ELEGIDA = "CIV - Cámara Nacional de Apelaciones en lo Civil"
PARTE_ID = "formPublica:nomIntervParte"
BOTON_ID = "formPublica:buscarPorParteButton"
BOTON_SIGUIENTE_ID = "j_idt119:j_idt200:j_idt207"
TABLA_EXPEDIENTES_ID = '/html/body/div[1]/div[2]/div[1]/div[2]/div/div/div/form/table/tbody/tr/td[1]'
SITUACION_ID = '/html/body/div[1]/div[2]/div[1]/div[2]/div/div/div/form/table/tbody/tr/td[4]'
ULTIMA_ACT_ID = '/html/body/div[1]/div[2]/div[1]/div[2]/div/div/div/form/table/tbody/tr/td[5]'
BOTON_OJO_ID = "/html/body/div[1]/div[2]/div[1]/div[2]/div/div/div/form/table/tbody/tr/td[6]/a"
TIEMPO_ESPERA_CAPTCHA = 60
TIEMPO_ESPERA_NUEVA_ACCION = 2
TIEMPO_ESPERA_NUEVOS_EXPEDIENTES = 60
CANTIDAD_EXPEDIENTES  = 10
SITUCAION_REQUERIDA = ["EN DESPACHO", "EN LETRA", "GIRO"]
EXPEDIENTES_FINALES = []
FECHA_LIMITE = pd.to_datetime("31/12/2019")
driver = webdriver.Chrome(ChromeDriverManager().install()) 

# Implementacion

def main():
    # Abro Excel
    
    if os.path.exists('archivo_expedientes.xlsx'):
        new_wb = load_workbook('archivo_expedientes.xlsx')
    else:
        new_wb = Workbook()
        EXPEDIENTES_FINALES.append(["Expediente", "Jurisdicción", "Situación Actual", "Carátula", "Actores", "Demandados"])     

    archivo_expedientes = new_wb.active


    # Cargo pagina
    driver.get(WEB_PAGE)


    # Nos movemos a buscar por parte (id: formPublica:porParte:header:inactive)
    busqueda_por_parte = driver.find_element_by_id(BUSQUEDA_ID)
    busqueda_por_parte.click()
    time.sleep(TIEMPO_ESPERA_NUEVA_ACCION)

    # Seleccionamos Jurisdiccion ( "CIV – Cámara Nación de Apelaciones en lo Civil")
    jurisdiccion = driver.find_element_by_id(SELECT_ID)
    jurisdiccion_elegida = Select(jurisdiccion)
    jurisdiccion_elegida.select_by_visible_text(JURISDICCION_ELEGIDA)

    # Escribimos en PARTE "SOCIEDAD"
    parte = driver.find_element_by_name(PARTE_ID)
    parte.send_keys(VALOR_BUSQUEDA)

    # Esperamos 30 segs para la CAPTCHA
    time.sleep(TIEMPO_ESPERA_CAPTCHA)

    # Clickeamos boton consultar
    boton_consultar = driver.find_element_by_id(BOTON_ID)
    boton_consultar.click()

    # Iteramos sobre los expedientes hasta obtener 100
    contador = 0
    while (contador < CANTIDAD_EXPEDIENTES): 
        expedientes = driver.find_elements_by_xpath(TABLA_EXPEDIENTES_ID) 
        for x in range(0, len(expedientes)-1):
            expediente_nuevo = []
            situaciones = driver.find_elements_by_xpath(SITUACION_ID)
            ultima_actividades = driver.find_elements_by_xpath(ULTIMA_ACT_ID)
            botones_ojo = driver.find_elements_by_xpath(BOTON_OJO_ID)  
            situacion = situaciones[x].text
            ult_act = pd.to_datetime(ultima_actividades[x].text)
            if ((situacion in SITUCAION_REQUERIDA) and  (FECHA_LIMITE < ult_act) and (contador < CANTIDAD_EXPEDIENTES)):
                try:
                    boton_ojo = botones_ojo[x]
                    boton_ojo.click()
                    time.sleep(1)
                    expediente_nuevo.append(driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[2]/div/div/fieldset/div/div[1]/div/div/div[2]/span').text)
                    expediente_nuevo.append(driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[2]/div/div/fieldset/div/div[2]/div/div/div[2]/span').text)
                    expediente_nuevo.append(driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[2]/div/div/fieldset/div/div[4]/div/div/div[2]/span').text)
                    expediente_nuevo.append(driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[2]/div/div/fieldset/div/div[5]/div/div/div[2]/span').text)
                    tab_intervinientes = driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[3]/div/div[1]/table/tbody/tr/td[6]')
                    tab_intervinientes.click()
                    time.sleep(1)
                    expediente_nuevo.append(driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[3]/div/div[4]/div/div/table/tbody[1]/tr/td[2]/span[2]').text)
                    expediente_nuevo.append(driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[3]/div/div[4]/div/div/table/tbody[3]/tr/td[2]/span[2]').text)
                    
                    EXPEDIENTES_FINALES.append(expediente_nuevo)
                    boton_volver = driver.find_element_by_xpath('/html/body/div[1]/div[2]/form/div/div[1]/div/div/a')
                    boton_volver.click()
                    contador = contador + 1
                    print("apendeo")
                    time.sleep(5)
                except:
                    print("error")
                    continue
        print(contador)       
        if (contador >= CANTIDAD_EXPEDIENTES):
            break
        boton_siguiente = driver.find_element_by_id(BOTON_SIGUIENTE_ID)
        boton_siguiente.click()
        time.sleep(TIEMPO_ESPERA_NUEVOS_EXPEDIENTES)
    
    for expediente in EXPEDIENTES_FINALES:
        archivo_expedientes.append(expediente)
    new_wb.save("archivo_expedientes.xlsx")
    driver.quit()

main()

